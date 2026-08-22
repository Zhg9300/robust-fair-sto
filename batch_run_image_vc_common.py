"""Shared matrix, scheduler, and result writer for V.C image experiments."""

import argparse
import csv
import itertools
import json
import os
import queue
import re
import shlex
import subprocess
import sys
import threading
import time
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


PROJECT_ROOT = Path(__file__).resolve().parent
ROBUST_AGGREGATORS = (
    "cwm",
    "cwtm",
    "median",
    "krum",
    "mda",
    "faba",
    "nbs",
)
AGGREGATORS = ("mean",) + ROBUST_AGGREGATORS
ATTACK_MODES = ("gaussian", "sign_flip", "adaptive_copying")
FLOAT_PATTERN = r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?"

RUN_PARAM_KEYS = (
    "seed",
    "device",
    "module",
    "algorithm",
    "dataloader",
    "N",
    "C",
    "B",
    "micro_batch_size",
    "data_device",
    "partition_seed",
    "min_client_samples",
    "NC",
    "balance",
    "Diralpha",
    "R",
    "E",
    "test_interval",
    "sgd_step",
    "lr",
    "decay",
    "momentum",
    "weight_decay",
    "q",
    "qffl_update_rule",
    "lam",
    "gradient_aggregator",
    "gradient_aggregator_f",
    "attack_mode",
    "dishonest_num",
    "byzantine_ids",
    "attack_start_round",
    "attack_end_round",
    "attack_seed",
    "attack_scale",
    "loss_bias",
    "attack_target_clients",
    "copy_loss",
    "copy_gradient",
    "evaluation_excluded_ids",
)

TABLE_FIELDS = (
    "experiment_no",
    "experiment_id",
    "dataset",
    "partition",
    "scenario",
    "seed",
    "partition_seed",
    "attack_seed",
    "algorithm",
    "module",
    "dataloader",
    "N",
    "NC",
    "Diralpha",
    "C",
    "B",
    "R",
    "E",
    "sgd_step",
    "lr",
    "decay",
    "momentum",
    "weight_decay",
    "q",
    "qffl_update_rule",
    "lam",
    "attack_mode",
    "attack_scale",
    "dishonest_num",
    "byzantine_ids",
    "attack_start_round",
    "attack_end_round",
    "attack_target_clients",
    "copy_loss",
    "copy_gradient",
    "evaluation_excluded_ids",
    "gradient_aggregator",
    "gradient_aggregator_f",
    "accuracy_mean",
    "accuracy_variance",
    "accuracy_min",
    "accuracy_max",
    "loss_mean",
    "loss_variance",
    "loss_min",
    "loss_max",
    "accuracy_list",
    "loss_list",
    "status",
    "return_code",
    "assigned_gpu",
    "gpu_slot",
    "device",
    "start_time",
    "end_time",
    "duration_seconds",
    "log_file",
    "command",
    "error",
)


@dataclass(frozen=True)
class DatasetSpec:
    slug: str
    display_name: str
    module: str
    rounds: int
    pathological_loader: str
    dirichlet_loader: str


def _float_token(value):
    return f"{float(value):g}".replace("-", "m").replace(".", "p")


def _algorithm_points():
    points = []
    for lr, q in itertools.product((0.1, 0.2, 0.4, 0.8), (0.1, 0.5, 1.0)):
        points.append({
            "algorithm": "qFedAvg",
            "lr": lr,
            "q": q,
            "qffl_update_rule": "objective_gradient",
            "decay": 1.0,
        })
    for lr in (0.3, 0.4, 0.5):
        points.append({
            "algorithm": "DRFL",
            "lr": lr,
            "decay": 0.999,
        })
    for lam, lr in itertools.product((0.05, 0.1, 0.4, 0.8), (0.01, 0.05, 0.1)):
        points.append({
            "algorithm": "AFL",
            "lam": lam,
            "lr": lr,
            "decay": 0.999,
        })
    return points


def _partition_points(spec):
    return (
        ({
            "_partition": "pathological_nc1",
            "dataloader": spec.pathological_loader,
            "NC": 1,
            "Diralpha": None,
        }),
        ({
            "_partition": "dirichlet_alpha0.5",
            "dataloader": spec.dirichlet_loader,
            "NC": 1,
            "Diralpha": 0.5,
        }),
    )


def _base_config(spec, seed, partition, algorithm_point):
    config = {
        "_dataset": spec.display_name,
        "seed": int(seed),
        "partition_seed": int(seed),
        "attack_seed": int(seed),
        "module": spec.module,
        "N": 10,
        "C": 1.0,
        "B": 64,
        "micro_batch_size": 0,
        "data_device": "model",
        "min_client_samples": "auto",
        "balance": True,
        "R": spec.rounds,
        "E": 1,
        "test_interval": 50,
        "sgd_step": True,
        "momentum": 0.0,
        "weight_decay": 5e-4,
        "attack_start_round": 1,
        "attack_end_round": "None",
        "loss_bias": 0.0,
        "evaluation_excluded_ids": "8,9",
        **partition,
        **algorithm_point,
    }
    return config


def _experiment_id(config):
    algorithm_bits = [
        config["algorithm"].lower(),
        f"lr{_float_token(config['lr'])}",
    ]
    if config["algorithm"] == "qFedAvg":
        algorithm_bits.append(f"q{_float_token(config['q'])}")
    elif config["algorithm"] == "AFL":
        algorithm_bits.append(f"lam{_float_token(config['lam'])}")
    return "_".join((
        config["_partition"],
        *algorithm_bits,
        config["_scenario"],
        config["gradient_aggregator"],
        f"seed{config['seed']}",
    ))


def build_experiments(spec, seeds=(1,)):
    """Build the complete clean/attack matrix for one image dataset."""
    experiments = []
    for seed in seeds:
        for partition in _partition_points(spec):
            for algorithm_point in _algorithm_points():
                base = _base_config(spec, seed, partition, algorithm_point)

                clean = {
                    **base,
                    "_scenario": "clean",
                    "gradient_aggregator": "mean",
                    "gradient_aggregator_f": 0,
                    "attack_mode": "None",
                    "attack_scale": 1.0,
                    "dishonest_num": 0,
                    "byzantine_ids": None,
                    "attack_target_clients": None,
                    "copy_loss": False,
                    "copy_gradient": False,
                }
                clean["_experiment_id"] = _experiment_id(clean)
                experiments.append(clean)

                for attack_mode in ATTACK_MODES:
                    for aggregator in AGGREGATORS:
                        attacked = {
                            **base,
                            "_scenario": f"{attack_mode}_attack",
                            "gradient_aggregator": aggregator,
                            "gradient_aggregator_f": 2,
                            "attack_mode": attack_mode,
                            "attack_scale": 5.0 if attack_mode == "gaussian" else 1.0,
                            "dishonest_num": 2,
                            "byzantine_ids": "8,9",
                            "attack_target_clients": (
                                "0" if attack_mode == "adaptive_copying" else None
                            ),
                            "copy_loss": attack_mode == "adaptive_copying",
                            "copy_gradient": attack_mode == "adaptive_copying",
                        }
                        attacked["_experiment_id"] = _experiment_id(attacked)
                        experiments.append(attacked)
    return experiments


def build_command(config):
    command = [sys.executable, str(PROJECT_ROOT / "run.py")]
    for key in RUN_PARAM_KEYS:
        value = config.get(key)
        if value is not None:
            command.extend((f"--{key}", str(value)))
    return command


def stringify_command(command):
    return " ".join(shlex.quote(part) for part in command)


def _last_groups(pattern, text):
    matches = list(re.finditer(pattern, text, flags=re.MULTILINE))
    return matches[-1].groups() if matches else None


def _parse_number_list(text, label):
    matches = re.findall(
        re.escape(label) + r"\s*\[(.*?)\]",
        text,
        flags=re.DOTALL,
    )
    if not matches:
        return None
    return [float(value) for value in re.findall(FLOAT_PATTERN, matches[-1])]


def parse_metrics(text):
    metrics = {
        "accuracy_mean": None,
        "accuracy_variance": None,
        "accuracy_min": None,
        "accuracy_max": None,
        "loss_mean": None,
        "loss_variance": None,
        "loss_min": None,
        "loss_max": None,
        "accuracy_list": None,
        "loss_list": None,
    }
    accuracy = _last_groups(
        r"^Average:\s*(" + FLOAT_PATTERN + r")\.\s*"
        r"Variance:\s*(" + FLOAT_PATTERN + r")\.\s*"
        r"Min:\s*(" + FLOAT_PATTERN + r")\.\s*"
        r"Max:\s*(" + FLOAT_PATTERN + r")\s*$",
        text,
    )
    if accuracy:
        for key, value in zip(
            ("accuracy_mean", "accuracy_variance", "accuracy_min", "accuracy_max"),
            accuracy,
        ):
            metrics[key] = float(value)

    loss = _last_groups(
        r"^Loss Average:\s*(" + FLOAT_PATTERN + r")\.\s*"
        r"Loss Variance:\s*(" + FLOAT_PATTERN + r")\.\s*"
        r"Loss Min:\s*(" + FLOAT_PATTERN + r")\.\s*"
        r"Loss Max:\s*(" + FLOAT_PATTERN + r")\s*$",
        text,
    )
    if loss:
        for key, value in zip(
            ("loss_mean", "loss_variance", "loss_min", "loss_max"),
            loss,
        ):
            metrics[key] = float(value)

    accuracy_list = _parse_number_list(text, "Test Acc List:")
    loss_list = _parse_number_list(text, "Test Loss List:")
    metrics["accuracy_list"] = (
        json.dumps(accuracy_list) if accuracy_list is not None else None
    )
    metrics["loss_list"] = json.dumps(loss_list) if loss_list is not None else None
    return metrics


class ResultTable:
    """Thread-safe incremental Excel writer with a CSV fallback."""

    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.lock = threading.Lock()
        self.workbook = None
        self.csv_handle = None
        self.csv_writer = None
        if HAS_OPENPYXL:
            self.path = self.output_dir / "results.xlsx"
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
        else:
            self.path = self.output_dir / "results.csv"
            self.csv_handle = self.path.open("w", newline="", encoding="utf-8")
            self.csv_writer = csv.DictWriter(
                self.csv_handle,
                fieldnames=("sheet",) + TABLE_FIELDS,
            )
            self.csv_writer.writeheader()

    @staticmethod
    def _sheet_name(row):
        name = re.sub(r"[\\/*?:\[\]]", "_", str(row.get("algorithm", "Unknown")))
        return (name.strip() or "Unknown")[:31]

    def _worksheet(self, name):
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        worksheet = self.workbook.create_sheet(name)
        worksheet.append(TABLE_FIELDS)
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, len(TABLE_FIELDS)).coordinate}"
        return worksheet

    def write(self, row, flush=True):
        normalized = {field: row.get(field, "") for field in TABLE_FIELDS}
        sheet = self._sheet_name(row)
        with self.lock:
            if self.workbook is not None:
                worksheet = self._worksheet(sheet)
                worksheet.append([normalized[field] for field in TABLE_FIELDS])
                if normalized["status"] == "FAILED":
                    failure_fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid",
                    )
                    for cell in worksheet[worksheet.max_row]:
                        cell.fill = failure_fill
                if flush:
                    self.workbook.save(self.path)
            else:
                self.csv_writer.writerow({"sheet": sheet, **normalized})
                if flush:
                    self.csv_handle.flush()

    def close(self):
        with self.lock:
            if self.workbook is not None:
                if not self.workbook.sheetnames:
                    self._worksheet("Results")
                self.workbook.save(self.path)
            elif self.csv_handle is not None:
                self.csv_handle.flush()
                self.csv_handle.close()


def query_gpu_info():
    try:
        completed = subprocess.run(
            (
                "nvidia-smi",
                "--query-gpu=index,memory.used,memory.total",
                "--format=csv,noheader,nounits",
            ),
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
    except (OSError, subprocess.SubprocessError):
        return {}
    info = {}
    for line in completed.stdout.splitlines():
        parts = [part.strip() for part in line.split(",")]
        if len(parts) != 3:
            continue
        try:
            gpu_id, used, total = (int(part) for part in parts)
        except ValueError:
            continue
        info[gpu_id] = {"used": used, "total": total}
    return info


class DeviceSlotManager:
    def __init__(self, devices=None, max_per_device=2):
        available = sorted(query_gpu_info())
        self.devices = list(devices) if devices is not None else (available or [-1])
        if not self.devices:
            raise ValueError("At least one execution device is required.")
        self.max_per_device = int(max_per_device)
        if self.max_per_device <= 0:
            raise ValueError("max_per_gpu must be positive.")
        self.running = {device: 0 for device in self.devices}
        self.condition = threading.Condition()

    def acquire(self):
        with self.condition:
            while True:
                gpu_info = query_gpu_info()
                candidates = []
                for device in self.devices:
                    count = self.running[device]
                    if count >= self.max_per_device:
                        continue
                    used = gpu_info.get(device, {}).get("used", 0 if device == -1 else 10**9)
                    candidates.append((count, used, device))
                if candidates:
                    _, _, device = min(candidates)
                    self.running[device] += 1
                    return device, self.running[device]
                self.condition.wait(timeout=5)

    def release(self, device):
        with self.condition:
            self.running[device] = max(0, self.running[device] - 1)
            self.condition.notify_all()


def subprocess_environment():
    environment = os.environ.copy()
    conda_prefix = environment.get("CONDA_PREFIX")
    if conda_prefix:
        conda_library = str(Path(conda_prefix) / "lib")
        current = environment.get("LD_LIBRARY_PATH", "")
        environment["LD_LIBRARY_PATH"] = (
            conda_library if not current else conda_library + os.pathsep + current
        )
    return environment


def _row_base(experiment_no, config):
    row = {field: config.get(field, "") for field in TABLE_FIELDS}
    row.update({
        "experiment_no": experiment_no,
        "experiment_id": config["_experiment_id"],
        "dataset": config["_dataset"],
        "partition": config["_partition"],
        "scenario": config["_scenario"],
    })
    return row


def _safe_log_name(experiment_no, experiment_id):
    safe_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", experiment_id)
    return f"exp_{experiment_no:04d}_{safe_id}.log"


def _tail(text, length=500):
    compact = " ; ".join(line.strip() for line in text.splitlines() if line.strip())
    return compact[-length:]


def run_experiment(experiment_no, config, device_manager, table, output_dir):
    device, gpu_slot = device_manager.acquire()
    local = deepcopy(config)
    local["device"] = device
    command = build_command(local)
    command_text = stringify_command(command)
    log_path = Path(output_dir) / "logs" / _safe_log_name(
        experiment_no,
        local["_experiment_id"],
    )
    started_at = datetime.now()
    started = time.perf_counter()
    status = "FAILED"
    return_code = ""
    error = ""
    metrics = parse_metrics("")
    stdout = ""
    stderr = ""
    try:
        completed = subprocess.run(
            command,
            cwd=PROJECT_ROOT,
            env=subprocess_environment(),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
        return_code = completed.returncode
        stdout = completed.stdout or ""
        stderr = completed.stderr or ""
        metrics = parse_metrics(stdout + "\n" + stderr)
        parsed = metrics["accuracy_mean"] is not None and metrics["loss_mean"] is not None
        status = "SUCCESS" if completed.returncode == 0 and parsed else "FAILED"
        if completed.returncode != 0:
            error = _tail(stderr or stdout)
        elif not parsed:
            error = "Process completed but final classification metrics were not parsed."
    except Exception as exc:  # preserve the remaining batch when one task fails
        error = str(exc)[:500]
    finally:
        ended_at = datetime.now()
        duration = time.perf_counter() - started
        device_manager.release(device)

    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", encoding="utf-8") as handle:
        handle.write(f"Command: {command_text}\n")
        handle.write(f"Assigned device: {device}\n")
        handle.write(f"GPU slot: {gpu_slot}\n")
        handle.write(f"Return code: {return_code}\n")
        handle.write("=" * 32 + " STDOUT " + "=" * 32 + "\n")
        handle.write(stdout)
        handle.write("\n" + "=" * 32 + " STDERR " + "=" * 32 + "\n")
        handle.write(stderr)
        if error:
            handle.write("\n[ERROR]\n" + error + "\n")

    row = _row_base(experiment_no, local)
    row.update(metrics)
    row.update({
        "status": status,
        "return_code": return_code,
        "assigned_gpu": device,
        "gpu_slot": gpu_slot,
        "device": device,
        "start_time": started_at.isoformat(timespec="seconds"),
        "end_time": ended_at.isoformat(timespec="seconds"),
        "duration_seconds": round(duration, 3),
        "log_file": str(log_path),
        "command": command_text,
        "error": error,
    })
    table.write(row)
    print(
        f"[{status}] {experiment_no}: {local['_experiment_id']} "
        f"device={device} duration={duration:.1f}s",
        flush=True,
    )
    return status


def parse_seeds(value):
    try:
        seeds = tuple(int(item.strip()) for item in value.split(",") if item.strip())
    except ValueError as exc:
        raise argparse.ArgumentTypeError("seeds must be comma-separated integers") from exc
    if not seeds:
        raise argparse.ArgumentTypeError("at least one seed is required")
    if len(seeds) != len(set(seeds)):
        raise argparse.ArgumentTypeError("seeds must not contain duplicates")
    return seeds


def parse_devices(value):
    try:
        devices = tuple(int(item.strip()) for item in value.split(",") if item.strip())
    except ValueError as exc:
        raise argparse.ArgumentTypeError("gpus must be comma-separated integers") from exc
    if not devices:
        raise argparse.ArgumentTypeError("at least one GPU id is required")
    if len(devices) != len(set(devices)):
        raise argparse.ArgumentTypeError("GPU ids must not contain duplicates")
    return devices


def read_args(spec, argv=None):
    parser = argparse.ArgumentParser(
        description=f"Run the {spec.display_name} V.C tuning matrix."
    )
    parser.add_argument("--gpus", type=parse_devices, default=None)
    parser.add_argument("--max-per-gpu", type=int, default=2)
    parser.add_argument("--max-workers", type=int, default=None)
    parser.add_argument("--seeds", type=parse_seeds, default=(1,))
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)
    if args.max_per_gpu <= 0:
        parser.error("--max-per-gpu must be positive")
    if args.max_workers is not None and args.max_workers <= 0:
        parser.error("--max-workers must be positive")
    return args


def _resolve_output_dir(spec, requested):
    if requested is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = PROJECT_ROOT / "batch_results" / f"{spec.slug}_vc_{timestamp}"
    else:
        output_dir = requested.expanduser().resolve()
        if output_dir.exists() and any(output_dir.iterdir()):
            raise RuntimeError(f"output directory is not empty: {output_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _dry_run(experiments, devices, table, output_dir):
    commands_path = output_dir / "commands.txt"
    with commands_path.open("w", encoding="utf-8") as command_file:
        for experiment_no, config in enumerate(experiments, 1):
            local = deepcopy(config)
            local["device"] = devices[(experiment_no - 1) % len(devices)]
            command_text = stringify_command(build_command(local))
            command_file.write(command_text + "\n")
            row = _row_base(experiment_no, local)
            row.update({
                "status": "DRY-RUN",
                "assigned_gpu": local["device"],
                "device": local["device"],
                "command": command_text,
            })
            table.write(row, flush=False)
    table.close()
    print(f"[DRY-RUN] wrote {len(experiments)} commands to {commands_path}")
    print(f"[DRY-RUN] result manifest: {table.path}")


def run_cli(spec, argv=None):
    args = read_args(spec, argv)
    experiments = build_experiments(spec, args.seeds)
    output_dir = _resolve_output_dir(spec, args.output_dir)
    table = ResultTable(output_dir)
    available = sorted(query_gpu_info())
    devices = list(args.gpus) if args.gpus is not None else (available or [-1])

    clean_count = sum(config["_scenario"] == "clean" for config in experiments)
    print(f"Dataset: {spec.display_name}")
    print(f"Experiments: {len(experiments)} (clean={clean_count}, attacked={len(experiments) - clean_count})")
    print(f"Devices: {devices}")
    print(f"Output: {output_dir}")
    if args.dry_run:
        _dry_run(experiments, devices, table, output_dir)
        return 0

    manager = DeviceSlotManager(devices, args.max_per_gpu)
    max_workers = args.max_workers or max(1, len(devices) * args.max_per_gpu)
    tasks = queue.Queue()
    for experiment_no, config in enumerate(experiments, 1):
        tasks.put((experiment_no, config))
    counts = {"SUCCESS": 0, "FAILED": 0}
    counts_lock = threading.Lock()

    def worker():
        while True:
            try:
                experiment_no, config = tasks.get_nowait()
            except queue.Empty:
                return
            try:
                status = run_experiment(
                    experiment_no,
                    config,
                    manager,
                    table,
                    output_dir,
                )
                with counts_lock:
                    counts[status] = counts.get(status, 0) + 1
            finally:
                tasks.task_done()

    workers = [threading.Thread(target=worker, daemon=True) for _ in range(max_workers)]
    for worker_thread in workers:
        worker_thread.start()
    for worker_thread in workers:
        worker_thread.join()
    table.close()
    print(
        f"Completed: success={counts['SUCCESS']} failed={counts['FAILED']} "
        f"table={table.path}"
    )
    return 1 if counts["FAILED"] else 0

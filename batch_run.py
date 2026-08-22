"""
批量实验运行脚本。

支持多 GPU 并发调度、参数网格展开、Excel/CSV 结果写入和
``--dry-run``。在 ``get_experiment_groups()`` 中维护固定参数与遍历参数。
"""

import argparse
import csv
import itertools
import os
import queue
import re
import shlex
import subprocess
import threading
from copy import deepcopy
from datetime import datetime
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl not installed. Results will be saved as CSV instead.")
    print("         Install with: pip install openpyxl")


# ============================================================
# 输出表头
# ============================================================

PARAM_KEYS = [
    'seed', 'device', 'module', 'algorithm', 'dataloader',
    'N', 'B', 'micro_batch_size', 'data_device', 'partition_seed',
    'min_client_samples', 'NC', 'R', 'E', 'lr', 'decay', 'momentum', 'weight_decay',
    'gradient_aggregator', 'gradient_aggregator_f', 'cc_tau', 'cc_iterations',
    'median_max_iterations', 'aggregator_tolerance',
    'prefer', 'rho', 'bm_kappa', 'bm_tau',
    'lam', 'q', 'qffl_update_rule', 'qffl_loss_mode', 'epsilon',
    'alpha', 'tau', 'theta', 's', 'pow',
    'test_interval', 'sgd_step', 'C', 'balance',
    'Diralpha', 'attack_mode', 'dishonest_num',
    'byzantine_ids', 'attack_start_round', 'attack_end_round',
    'attack_seed', 'attack_scale', 'alie_z', 'loss_bias', 'attack_target_clients',
    'copy_loss', 'copy_gradient', 'evaluation_excluded_ids',
]

# Result workbooks prioritize experiment identity, tunable parameters, attack
# settings, and metrics. Runtime bookkeeping is deliberately kept at the end.
RESULT_CONFIG_COLUMNS = [
    ('Algorithm', 'algorithm'), ('Module', 'module'), ('DataLoader', 'dataloader'),
    ('lr', 'lr'), ('decay', 'decay'), ('momentum', 'momentum'),
    ('weight_decay', 'weight_decay'),
    ('R', 'R'), ('E', 'E'), ('B', 'B'),
    ('micro_batch_size', 'micro_batch_size'), ('data_device', 'data_device'),
    ('partition_seed', 'partition_seed'), ('min_client_samples', 'min_client_samples'),
    ('C', 'C'),
    ('N', 'N'), ('NC', 'NC'),
    ('attack_mode', 'attack_mode'), ('dishonest_num', 'dishonest_num'),
    ('byzantine_ids', 'byzantine_ids'),
    ('attack_start_round', 'attack_start_round'),
    ('attack_end_round', 'attack_end_round'), ('attack_seed', 'attack_seed'),
    ('attack_scale', 'attack_scale'), ('alie_z', 'alie_z'),
    ('loss_bias', 'loss_bias'), ('attack_target_clients', 'attack_target_clients'),
    ('copy_loss', 'copy_loss'), ('copy_gradient', 'copy_gradient'),
    ('evaluation_excluded_ids', 'evaluation_excluded_ids'),
    ('prefer', 'prefer'), ('rho', 'rho'),
    ('bm_kappa', 'bm_kappa'), ('bm_tau', 'bm_tau'),
    ('lam', 'lam'), ('q', 'q'),
    ('qffl_update_rule', 'qffl_update_rule'),
    ('qffl_loss_mode', 'qffl_loss_mode'),
    ('epsilon', 'epsilon'), ('alpha', 'alpha'),
    ('tau', 'tau'),
    ('theta', 'theta'), ('s', 's'), ('pow', 'pow'),
    ('gradient_aggregator', 'gradient_aggregator'),
    ('gradient_aggregator_f', 'gradient_aggregator_f'),
    ('cc_tau', 'cc_tau'), ('cc_iterations', 'cc_iterations'),
    ('median_max_iterations', 'median_max_iterations'),
    ('aggregator_tolerance', 'aggregator_tolerance'),
    ('test_interval', 'test_interval'), ('sgd_step', 'sgd_step'),
    ('balance', 'balance'), ('Diralpha', 'Diralpha'), ('Seed', 'seed'),
]

RESULT_METRIC_COLUMNS = [
    ('Avg Acc', 'avg_acc'), ('Var Acc', 'var_acc'),
    ('Min Acc', 'min_acc'), ('Max Acc', 'max_acc'),
    ('Mean Loss', 'mean_loss'), ('Acc List', 'acc_list'),
]

EXCEL_HEADERS = (
    ['No.']
    + [header for header, _ in RESULT_METRIC_COLUMNS]
    + [header for header, _ in RESULT_CONFIG_COLUMNS]
    + [
        'Status', 'Assigned GPU', 'GPU Slot', 'Device',
        'Start Time', 'End Time', 'Duration (s)',
        'Log File', 'Command', 'Error Info',
    ]
)


# ============================================================
# GPU 管理
# ============================================================

def get_gpu_info() -> Dict[int, Dict[str, int]]:
    """查询所有 GPU 的显存信息。返回 {gpu_id: {'used': x, 'total': y}}。"""
    try:
        result = subprocess.run(
            [
                'nvidia-smi',
                '--query-gpu=index,memory.used,memory.total',
                '--format=csv,noheader,nounits',
            ],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
        gpu_info = {}
        for line in result.stdout.strip().split('\n'):
            if not line.strip():
                continue
            parts = [x.strip() for x in line.split(',')]
            if len(parts) < 3:
                continue
            gpu_id = int(parts[0])
            used_mb = int(parts[1])
            total_mb = int(parts[2])
            gpu_info[gpu_id] = {'used': used_mb, 'total': total_mb}
        return gpu_info
    except Exception as e:
        print(f"[WARNING] Failed to query GPU info: {e}")
        return {}


class GPUSlotManager:
    """管理 GPU 槽位。

    每张 GPU 可以同时跑多个实验（由 max_per_gpu 控制）。
    调度策略：
      1. 只从 allowed_gpus 中选；
      2. 先找未满载的 GPU；
      3. 在未满载 GPU 中，优先选当前运行数更少的；
      4. 如果运行数相同，再选 nvidia-smi 显存占用更低的。
    """

    def __init__(self, allowed_gpus: Optional[List[int]], max_per_gpu: int = 2):
        self.allowed_gpus = allowed_gpus
        self.max_per_gpu = max_per_gpu
        self.lock = threading.Lock()
        self.cond = threading.Condition(self.lock)
        self.running_count: Dict[int, int] = {}

        gpu_info = get_gpu_info()
        if allowed_gpus is None:
            if gpu_info:
                self.allowed_gpus = sorted(gpu_info.keys())
            else:
                self.allowed_gpus = [0]

        for gid in self.allowed_gpus:
            self.running_count[gid] = 0

    def _choose_best_gpu(self) -> Optional[int]:
        gpu_info = get_gpu_info()
        candidates = []
        for gid in self.allowed_gpus:
            current_jobs = self.running_count.get(gid, 0)
            if current_jobs >= self.max_per_gpu:
                continue
            used_mem = gpu_info.get(gid, {}).get('used', 10**9)
            total_mem = gpu_info.get(gid, {}).get('total', 0)
            candidates.append((current_jobs, used_mem, -total_mem, gid))

        if not candidates:
            return None
        candidates.sort()
        return candidates[0][-1]

    def acquire(self) -> Tuple[int, int]:
        """阻塞直到拿到一个 GPU 槽位。返回 (gpu_id, slot_index)。"""
        with self.cond:
            while True:
                gid = self._choose_best_gpu()
                if gid is not None:
                    self.running_count[gid] += 1
                    slot_index = self.running_count[gid]
                    return gid, slot_index
                self.cond.wait(timeout=5)

    def release(self, gpu_id: int) -> None:
        with self.cond:
            self.running_count[gpu_id] = max(0, self.running_count.get(gpu_id, 1) - 1)
            self.cond.notify_all()

    def snapshot(self) -> Dict[int, int]:
        with self.lock:
            return dict(self.running_count)


# ============================================================
# 输出解析
# ============================================================

def parse_experiment_output(output_text: str) -> Dict[str, Optional[object]]:
    """从实验输出中提取最终测试指标。"""
    results = {
        'avg_acc': None,
        'var_acc': None,
        'min_acc': None,
        'max_acc': None,
        'mean_loss': None,
        'acc_list': None,
    }

    acc_pattern = re.compile(
        r'Average:\s*([\d.]+)\.\s*Variance:\s*([\d.]+)\.\s*'
        r'Min:\s*([\d.]+)\.\s*Max:\s*([\d.]+)'
    )
    loss_pattern = re.compile(r'Mean Global Test loss:\s*([\d.]+)')
    acc_list_pattern = re.compile(r'Test Acc List:\s*\[(.+?)\]')

    for match in acc_pattern.finditer(output_text):
        results['avg_acc'] = float(match.group(1))
        results['var_acc'] = float(match.group(2))
        results['min_acc'] = float(match.group(3))
        results['max_acc'] = float(match.group(4))

    for match in loss_pattern.finditer(output_text):
        results['mean_loss'] = float(match.group(1))

    for match in acc_list_pattern.finditer(output_text):
        results['acc_list'] = match.group(1).strip()

    return results


# ============================================================
# Excel / CSV
# ============================================================

def style_header(ws):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def create_excel(filepath: str):
    """创建 Excel 文件。

    注意：openpyxl 保存时至少需要一个可见 sheet。
    因此先保留一个占位 sheet，等真正写入 algorithm sheet 后再删除。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '_INIT'
    style_header(ws)
    wb.save(filepath)
    return wb


def save_results_csv(filepath: str, all_rows: List[List[object]]) -> None:
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Algorithm Sheet'] + EXCEL_HEADERS)
        writer.writerows(all_rows)


class ResultWriter:
    """线程安全地写 Excel / CSV，并按 algorithm 分 sheet。"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.lock = threading.Lock()
        self.csv_rows: List[List[object]] = []

        if HAS_OPENPYXL and filepath.endswith('.xlsx'):
            self.wb = create_excel(filepath)
        else:
            self.wb = None

    def _sheet_name(self, result_dict: Dict[str, object]) -> str:
        algo = str(result_dict.get('config', {}).get('algorithm', 'Unknown'))
        algo = re.sub(r'[\\/*?:\[\]]', '_', algo).strip() or 'Unknown'
        return algo[:31]

    def _ensure_sheet(self, sheet_name: str):
        if sheet_name in self.wb.sheetnames:
            return self.wb[sheet_name]
        ws = self.wb.create_sheet(sheet_name)
        style_header(ws)

        # 删除初始化占位 sheet，避免触发 “At least one sheet must be visible”
        if '_INIT' in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            init_ws = self.wb['_INIT']
            self.wb.remove(init_ws)
        return ws

    def _build_row(self, result_dict: Dict[str, object]) -> List[object]:
        config = result_dict.get('config', {})
        row = [result_dict.get('exp_no', '')]
        row.extend(result_dict.get(key, '') for _, key in RESULT_METRIC_COLUMNS)
        row.extend(config.get(key, '') for _, key in RESULT_CONFIG_COLUMNS)
        row.extend([
            result_dict.get('status', ''),
            result_dict.get('assigned_gpu', ''),
            result_dict.get('gpu_slot', ''),
            config.get('device', ''),
            result_dict.get('start_time_str', ''),
            result_dict.get('end_time_str', ''),
            result_dict.get('duration', ''),
            result_dict.get('log_file', ''),
            result_dict.get('command', ''),
            result_dict.get('error_info', ''),
        ])
        if len(row) != len(EXCEL_HEADERS):
            raise RuntimeError(
                f'Result row has {len(row)} values for {len(EXCEL_HEADERS)} headers.'
            )
        return row

    def write(self, result_dict: Dict[str, object]) -> None:
        row = self._build_row(result_dict)
        sheet_name = self._sheet_name(result_dict)

        with self.lock:
            if self.wb is not None:
                ws = self._ensure_sheet(sheet_name)
                row_idx = ws.max_row + 1
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

                if result_dict.get('status') == 'FAILED':
                    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    for col_idx in range(1, len(row) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill

                self.wb.save(self.filepath)
            else:
                self.csv_rows.append([sheet_name] + row)
                save_results_csv(self.filepath, self.csv_rows)


# ============================================================
# 命令构建与参数展开
# ============================================================

def build_command_list(config: Dict[str, object]) -> List[str]:
    """将配置字典转换为 subprocess 可直接执行的参数列表。"""
    cmd = ['python', 'run.py']
    for key in PARAM_KEYS:
        if key in config and config[key] is not None:
            cmd.extend([f'--{key}', str(config[key])])
    return cmd


def stringify_command(cmd_list: List[str]) -> str:
    return ' '.join(shlex.quote(x) for x in cmd_list)


def expand_grid(base_config: Dict[str, object], grid: Dict[str, List[object]]) -> List[Dict[str, object]]:
    """对 grid 做笛卡尔积展开。"""
    if not grid:
        return [deepcopy(base_config)]

    keys = list(grid.keys())
    values = [grid[k] for k in keys]
    configs = []
    for combo in itertools.product(*values):
        conf = deepcopy(base_config)
        for k, v in zip(keys, combo):
            conf[k] = v
        configs.append(conf)
    return configs


# ============================================================
# 实验配置定义
# ============================================================

def get_default_config() -> Dict[str, object]:
    return {
        'seed': 1,
        'N': 10,
        'B': 16,
        'NC': 1,
        'E': 1,
        'decay': 0.999,
        'momentum': 0.9,
        'weight_decay': 0.0,
        'test_interval': 200,
    }


def get_experiment_groups() -> List[Dict[str, object]]:
    """在这里定义实验组。

    每个实验组包含：
        name: 实验组名称（仅用于阅读）
        base: 固定参数
        grid: 需要遍历的参数列表

    例如：
        'grid': {
            'lr': [0.01, 0.05, 0.1],
            'B': [16, 32],
            'seed': [1, 2, 3],
        }
    则会自动展开为 3 x 2 x 3 = 18 个实验。
    """
    base = get_default_config()
    groups = [
        {
            'name': 'fashion_test',
            'base': {
                **base,
                'module': 'MLP',
                'dataloader': 'DataLoader_fashion_pat',
                'R': 2001,
            },
            'grid': {
                'algorithm': ['FedAvg', 'DRFL'],
                'lr': [0.1, 0.2, 0.3],
                'B': ['full'],
            },
        },
        {
            'name': '2',
            'base': {
                **base,
                'module': 'CNN',   
                'sgd_step': 'True',             
                'R': 4001,
            },
            'grid': {
                'dataloader': ['DataLoader_cifar10_pat','DataLoader_cifar10_dir'],
                'algorithm': ['FedAvg'],
                'lr': [0.25,0.3,0.35,0.4,0.45],
                'B': [256, 128, 64, 32, 16],
            },
        },
        {
            'name': '2',
            'base': {
                **base,
                'module': 'MLP',   
                'sgd_step': 'True',             
                'R': 3001,
            },
            'grid': {
                'dataloader': ['DataLoader_fashion_pat','DataLoader_fashion_dir'],
                'algorithm': ['FedAvg'],
                'lr': [0.3,0.4,0.5,0.6],
                'B': [256, 128, 64, 32, 16],
            },
        },
        {
            'name': '1',
            'base': {
                **base,
                'module': 'CNN',
                'dataloader': 'DataLoader_cifar10_dir',
                'R': 3001,
            },
            'grid': {
                'algorithm': ['FedAvg', 'DRFL'],
                'lr': [0.2, 0.3, 0.4, 0.5, 0.6],
                'B': ['full'],
            },
        },
        {
            'name': '1',
            'base': {
                **base,
                'module': 'CNN',
                'dataloader': 'DataLoader_cifar10_dir',
                'R': 3001,
            },
            'grid': {
                'algorithm': ['AdaFed'],
                'lr': [3,4,5,6],
                'pow': [1,2,3],
                'B': ['full'],
            },
        },
        {
            'name': '1',
            'base': {
                **base,
                'module': 'CNN',
                'dataloader': 'DataLoader_cifar10_pat',
                'R': 3001,
            },
            'grid': {
                'algorithm': ['qFedAvg'],
                'lr': [0.5,1,1.5,2],
                'q': [0.05, 0.1, 0.2],
                'B': ['full'],
            },
        },
        {
            'name': 'q',
            'base': {
                **base,
                'module': 'CNN',
                'dataloader': 'DataLoader_cifar10_dir',
                'R': 3001,
            },
            'grid': {
                'algorithm': ['qFedAvg'],
                'lr': [0.5,1,1.5,2],
                'q': [0.05, 0.1, 0.15, 0.2],
                'B': ['full'],
            },
        },
        {
            'name': '2',
            'base': {
                **base,
                'module': 'CNN',
                'dataloader': 'DataLoader_cifar10_dir',
                'R': 3001,
            },
            'grid': {
                'algorithm': ['FedMGDA_plus'],
                'lr': [3,5,7,9],
                'epsilon': [0.6,0.8,1],
                'B': ['full'],
            },
        },
    ]

    return groups


def get_experiments() -> List[Dict[str, object]]:
    """将所有实验组展开成实验列表。"""
    groups = get_experiment_groups()
    experiments = []
    for group in groups:
        base = deepcopy(group.get('base', {}))
        grid = deepcopy(group.get('grid', {}))
        expanded = expand_grid(base, grid)
        for idx, conf in enumerate(expanded, 1):
            conf['_group_name'] = group.get('name', 'group')
            conf['_group_index'] = idx
            experiments.append(conf)
    return experiments


# ============================================================
# 单实验执行
# ============================================================

def tail_text(text: str, lines: int = 6, max_len: int = 300) -> str:
    arr = [x for x in text.strip().split('\n') if x.strip()]
    if not arr:
        return ''
    s = ' ; '.join(arr[-lines:])
    return s[:max_len]


def run_single_experiment(
    exp_no: int,
    config: Dict[str, object],
    gpu_manager: GPUSlotManager,
    result_writer: ResultWriter,
    results_dir: str,
    dry_run: bool = False,
) -> Dict[str, object]:
    """执行单个实验。"""
    gpu_id, gpu_slot = gpu_manager.acquire()
    start_time = datetime.now()

    local_config = deepcopy(config)
    local_config['device'] = gpu_id
    cmd_list = build_command_list(local_config)
    cmd_str = stringify_command(cmd_list)

    log_dir = os.path.join(results_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(
        log_dir,
        f"exp_{exp_no:04d}_gpu{gpu_id}_slot{gpu_slot}_{start_time.strftime('%m%d_%H%M%S')}.log"
    )

    print(f"[START] Exp {exp_no}: GPU {gpu_id} Slot {gpu_slot}")
    print(f"        {cmd_str}")

    status = 'SUCCESS'
    error_info = ''
    parsed = {
        'avg_acc': '',
        'var_acc': '',
        'min_acc': '',
        'max_acc': '',
        'mean_loss': '',
        'acc_list': '',
    }

    try:
        if dry_run:
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write('[DRY RUN]\n')
                f.write(cmd_str + '\n')
            status = 'DRY-RUN'
        else:
            proc = subprocess.run(
                cmd_list,
                capture_output=True,
                text=True,
                timeout=None,
                check=False,
            )

            output = (proc.stdout or '') + '\n' + (proc.stderr or '')
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write(f"Command: {cmd_str}\n")
                f.write(f"Assigned GPU: {gpu_id}\n")
                f.write(f"GPU Slot: {gpu_slot}\n")
                f.write(f"Return code: {proc.returncode}\n")
                f.write('=' * 40 + ' STDOUT ' + '=' * 40 + '\n')
                f.write(proc.stdout or '')
                f.write('\n' + '=' * 40 + ' STDERR ' + '=' * 40 + '\n')
                f.write(proc.stderr or '')

            if proc.returncode != 0:
                status = 'FAILED'
                error_info = tail_text(proc.stderr or output)
            else:
                parsed = parse_experiment_output(output)
                if parsed.get('avg_acc') is None and parsed.get('mean_loss') is None:
                    # 运行成功但没有解析到结果，给一个弱提示
                    error_info = 'Process finished successfully, but no metrics were parsed from output.'

    except Exception as e:
        status = 'FAILED'
        error_info = str(e)[:300]
        try:
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write('\n[EXCEPTION]\n')
                f.write(error_info + '\n')
        except Exception:
            pass

    finally:
        end_time = datetime.now()
        duration = round((end_time - start_time).total_seconds(), 1)
        gpu_manager.release(gpu_id)

    result = {
        'exp_no': exp_no,
        'status': status,
        'start_time_str': start_time.strftime('%Y-%m-%d %H:%M:%S'),
        'end_time_str': end_time.strftime('%Y-%m-%d %H:%M:%S'),
        'duration': duration,
        'assigned_gpu': gpu_id,
        'gpu_slot': gpu_slot,
        'log_file': log_file,
        'command': cmd_str,
        'config': local_config,
        'avg_acc': parsed.get('avg_acc', ''),
        'var_acc': parsed.get('var_acc', ''),
        'min_acc': parsed.get('min_acc', ''),
        'max_acc': parsed.get('max_acc', ''),
        'mean_loss': parsed.get('mean_loss', ''),
        'acc_list': parsed.get('acc_list', ''),
        'error_info': error_info,
    }

    result_writer.write(result)

    if status == 'FAILED':
        print(f"[FAILED] Exp {exp_no}: GPU {gpu_id}, duration={duration}s")
        print(f"         Error: {error_info}")
    elif status == 'DRY-RUN':
        print(f"[DRY-RUN] Exp {exp_no} prepared on GPU {gpu_id}, slot {gpu_slot}")
    else:
        print(
            f"[DONE] Exp {exp_no}: GPU {gpu_id}, duration={duration}s, "
            f"avg_acc={result['avg_acc']}, loss={result['mean_loss']}"
        )

    return result


# ============================================================
# 主逻辑
# ============================================================

def parse_gpu_list(gpu_str: Optional[str]) -> Optional[List[int]]:
    if gpu_str is None:
        return None
    return [int(x.strip()) for x in gpu_str.split(',') if x.strip()]


def main():
    parser = argparse.ArgumentParser(description='Concurrent batch experiment runner')
    parser.add_argument('--gpus', type=str, default=None,
                        help='Comma-separated GPU IDs, e.g. "0,1,2,3"')
    parser.add_argument('--max-per-gpu', type=int, default=2,
                        help='Maximum concurrent experiments on one GPU')
    parser.add_argument('--max-workers', type=int, default=None,
                        help='Maximum total worker threads. Default = num_gpus * max_per_gpu')
    parser.add_argument('--dry-run', action='store_true',
                        help='Only print commands, do not execute')
    args = parser.parse_args()

    allowed_gpus = parse_gpu_list(args.gpus)
    experiments = get_experiments()
    total = len(experiments)

    initial_gpu_info = get_gpu_info()
    if allowed_gpus is None:
        gpu_pool = sorted(initial_gpu_info.keys()) if initial_gpu_info else [0]
    else:
        gpu_pool = allowed_gpus

    max_workers = args.max_workers or max(1, len(gpu_pool) * args.max_per_gpu)

    print('=' * 80)
    print('[CONFIG] Concurrent batch experiment runner')
    print(f'[CONFIG] Allowed GPUs: {gpu_pool}')
    print(f'[CONFIG] Max experiments per GPU: {args.max_per_gpu}')
    print(f'[CONFIG] Total worker threads: {max_workers}')
    print(f'[CONFIG] Total experiments: {total}')
    print(f'[CONFIG] Dry run: {args.dry_run}')
    print('=' * 80)

    timestamp = datetime.now().strftime('%m%d_%H%M%S')
    results_dir = 'batch_results'
    os.makedirs(results_dir, exist_ok=True)

    if HAS_OPENPYXL:
        results_file = os.path.join(results_dir, f'results_{timestamp}.xlsx')
    else:
        results_file = os.path.join(results_dir, f'results_{timestamp}.csv')

    print(f'[CONFIG] Results file: {results_file}')
    print('=' * 80)

    gpu_manager = GPUSlotManager(allowed_gpus=gpu_pool, max_per_gpu=args.max_per_gpu)
    result_writer = ResultWriter(results_file)

    task_queue: "queue.Queue[Tuple[int, Dict[str, object]]]" = queue.Queue()
    for idx, exp in enumerate(experiments, 1):
        task_queue.put((idx, exp))

    summary_lock = threading.Lock()
    summary = {'SUCCESS': 0, 'FAILED': 0, 'DRY-RUN': 0}

    def worker_loop(worker_id: int):
        while True:
            try:
                exp_no, exp_conf = task_queue.get_nowait()
            except queue.Empty:
                return

            try:
                result = run_single_experiment(
                    exp_no=exp_no,
                    config=exp_conf,
                    gpu_manager=gpu_manager,
                    result_writer=result_writer,
                    results_dir=results_dir,
                    dry_run=args.dry_run,
                )
                with summary_lock:
                    status = result.get('status', 'FAILED')
                    summary[status] = summary.get(status, 0) + 1
            finally:
                task_queue.task_done()

    workers = []
    for wid in range(max_workers):
        t = threading.Thread(target=worker_loop, args=(wid + 1,), daemon=True)
        t.start()
        workers.append(t)

    for t in workers:
        t.join()

    print('\n' + '=' * 80)
    print('[DONE] Batch completed!')
    print(f"  Total:   {total}")
    print(f"  Success: {summary.get('SUCCESS', 0)}")
    print(f"  Failed:  {summary.get('FAILED', 0)}")
    print(f"  DryRun:  {summary.get('DRY-RUN', 0)}")
    print(f"  Results: {results_file}")
    print('=' * 80)


if __name__ == '__main__':
    main()
